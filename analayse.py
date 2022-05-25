from lib import *
import openpyxl

class KeyWord:
    def __init__(self,student_answer,answer_key):
        self.answer_key = answer_key
        self.student_answer = student_answer
        self.metadata = ""
        self.parseChar = [',','.','!','“','”','(',')','{','}','[',']','"',"'"] 
        self.frequency = {} 
        self.validate() #function to validate the path
        self.raw_answer_key = [] #raw data from answer key
        self.raw_answer_std = [] #raw data form student answer
        self.key_dict = {} #to store the keywords in answer key
        self.mark_dict = {} #store the total mark for each question in answer key
        self.question_dict = {} #store the question code and question string
        self.answer_dict = {} #store the answer of the student with respect to question id
        self.mark_dict_eval = {} #student mark with respect to question 
        self.std_id = {} #student id and corresponding question id
        self.filehandler()
        

    def validate(self):
        split_path = os.path.splitext(self.student_answer)
        split_path_2 = os.path.splitext(self.answer_key)
        extension = split_path[-1]
        extension_2 = split_path_2[-1]
        if extension == ".xlsx" and extension_2 == ".xlsx":
            self.metadata = "excel_file"
        else:
            self.metadata = "incompatible"
    
    def filehandler(self):
        if self.metadata == "incompatible":
            self.metadata = "ERROR"
        elif self.metadata == "excel_file":
            self.cleanAnswerKey()
            self.cleanStudentAnswer()
            self.hitCounter()
            self.editMark()
        
    
    def cleanAnswerKey(self):
        wb_answer_key = openpyxl.load_workbook(self.answer_key)
        sheet = wb_answer_key.active
        for i in range(2, sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                obj_cell = sheet.cell(row=i, column=j)
                self.raw_answer_key.append(obj_cell.value)
        
            for i in range(0,len(self.raw_answer_key),5):
                self.question_dict[self.raw_answer_key[i]] = self.raw_answer_key[i+1]
                self.key_dict[self.raw_answer_key[i]] = self.raw_answer_key[i+3]
                self.mark_dict[self.raw_answer_key[i]] = self.raw_answer_key[i+4]
        wb_answer_key.close()

    def cleanStudentAnswer(self):
        wb_answer_std = openpyxl.load_workbook(self.student_answer)
        sheet = wb_answer_std.active
        for i in range(2, sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                obj_cell = sheet.cell(row=i, column=j)
                self.raw_answer_std.append(obj_cell.value)
            
        for i in range(0,len(self.raw_answer_std),5):
            self.std_id[self.raw_answer_std[i]] = self.raw_answer_std[i+1]
            self.answer_dict[self.raw_answer_std[i+1]] = self.raw_answer_std[i+3]
            self.mark_dict_eval[self.raw_answer_std[i+1]] = self.raw_answer_std[i+4]
        wb_answer_std.close

    def markCalculator(self, hitcount, keycount, i):
        percentage = (hitcount/keycount)*100
        mark = round((percentage/100)*self.mark_dict[i],2)
        self.mark_dict_eval[i] = mark


    def hitCounter(self):
        for i in self.answer_dict:
            if i in self.key_dict.keys():
                hit = 0
                keylist = self.key_dict[i].split(",")
                anslist = self.answer_dict[i].split()
                for word in anslist:
                    for char in self.parseChar:
                        word = word.rstrip(char) #right stripping
                        word = word.lstrip(char) #left stripping
                    if word in keylist:
                        if word in self.frequency:
                            self.frequency[word] += 1
                        else:
                            hit += 1
                            self.frequency[word] = 1
                self.markCalculator(hit,len(keylist),i)


    def editMark(self):
        wb_answer_std = openpyxl.load_workbook(self.student_answer)
        sheet = wb_answer_std.active
        for i in range(2, sheet.max_row+1):
            if sheet.cell(row=i,column=2).value in self.mark_dict_eval.keys():
                q_code = sheet.cell(row=i,column=2).value
                mark = str(self.mark_dict_eval[q_code])
                mark  += " out of " + str(self.mark_dict[q_code])
                sheet.cell(row=i,column=5).value = mark
                #sheet.cell(row=i,column=5).value = self.mark_dict_eval[q_code]

        wb_answer_std.save(self.student_answer)
